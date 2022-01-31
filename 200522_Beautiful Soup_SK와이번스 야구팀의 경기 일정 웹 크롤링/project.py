import requests   #웹 페이지 HTML 가져옴
from bs4 import BeautifulSoup   #HTML 파싱

for month in range(6, 11, 1):
    url = "https://sports.news.naver.com/kbaseball/schedule/index.nhn?date=20200616&month=0"
    url = url + str(month) + "&year=2020&teamCode="
    request = requests.get(url)
    html = request.text
    soup = BeautifulSoup(html, 'html.parser')
    print(url)


DateList = soup.findAll("td", {"rowspan" : "5"})   #경기 날짜 리스트, 경기 일정이 있는 날짜의 rowspan 값이 5, 없는 날은 1
HomeList = soup.findAll("span", {"class" : "team_rgt"})   #홈팀 리스트
AwayList = soup.findAll("span", {"class" : "team_lft"})   #어웨이팀 리스트
td_stadiumList = soup.findAll("span", {"class" : "td_stadium"})   #경기장 리스트
StadiumList = []
for number in range(len(DateList)):
    DateList[number] = DateList[number].find("span", {"class": "td_date"}).text
        
for number in range(len(HomeList)):
    HomeList[number] = HomeList[number].text
    AwayList[number] = AwayList[number].text

for number in range(1, len(td_stadiumList), 2):   # 짝수번째 값만 다시 저장, 중계 방송사와 경기장이 같은 속성 값을 갖고 있기 때문
    StadiumList.append(td_stadiumList[number].text)


from openpyxl import Workbook
wb = Workbook()
sheet1 = wb.active
file_name = "2020SKWyberns.xlsx"
sheet1.title = 'SK Wyverns'
row_index = [2]
for number in range(len(StadiumList)):
    if HomeList[number] == 'SK' or AwayList[number] == 'SK':
            sheet1.cell(row = row_index[0], column = 2).value = DateList[number//5]
            sheet1.cell(row = row_index[0], column = 3).value = AwayList[number]
            sheet1.cell(row = row_index[0], column = 4).value = HomeList[number]
            sheet1.cell(row = row_index[0], column = 5).value = StadiumList[number]
            row_index[0] += 1
wb.save(filename = file_name)
wb.close()


import pandas as pd
sk = pd.read_excel("2020SKWyberns.xlsx", encoding = 'euc-kr')
sk

sk.loc[30]

sk.set_index(keys = ['월', '날짜'])

Munhak = sk['구장'] == '문학'
sk[Munhak].set_index(keys = ['월', '날짜'])














